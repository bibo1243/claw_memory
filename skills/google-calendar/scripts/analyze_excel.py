import pandas as pd
import sys

def analyze_excel(file_path):
    try:
        # Attempt to read with different engines if openpyxl is not default
        # Since we might not have pandas installed in the environment, we'll try a raw python approach first 
        # or rely on pre-installed libraries.
        # But wait, I see previous attempts failed to install pandas.
        # Let's try to use 'openpyxl' if available, or just parse XML if it's XLSX (which is zipped XML).
        # Actually, let's try a simpler approach: convert xlsx to csv using a tool if available?
        # Or check what python libs ARE available.
        
        # Checking available modules
        import importlib.util
        if importlib.util.find_spec("openpyxl") is None:
             print("Error: openpyxl not installed.")
             return

        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
            
        # Basic Analysis (assuming headers in first row)
        if not data:
            print("File is empty.")
            return

        headers = data[0]
        rows = data[1:]
        
        print(f"Headers: {headers}")
        print(f"Total Rows: {len(rows)}")
        print("First 5 rows:")
        for r in rows[:5]:
            print(r)
            
    except Exception as e:
        print(f"Error analyzing file: {e}")

if __name__ == "__main__":
    analyze_excel(sys.argv[1])
